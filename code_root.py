import sys
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import speech_recognition as sr
from sympy import sympify
from decimal import Decimal, InvalidOperation 
import openpyxl

# Cấu hình UTF-8 để hỗ trợ tiếng Việt & tiếng Nga
sys.stdout.reconfigure(encoding='utf-8')

# Bảng chuyển đổi từ tiếng Việt và tiếng Nga sang ký hiệu toán học
TRANSLATION_DICT = {
    "cộng": "+", "trừ": "-", "nhân": "*", "chia": "/", "mũ": "**", "bằng": "=",
    "плюс": "+", "минус": "-", "умножить": "*", "разделить": "/", "степени": "**", "равно": "=",
    "x": "*",  # Thêm vào để chuyển "x" thành dấu nhân *
}

class VoiceCalculatorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Компьютер распознавания голоса")
        self.root.geometry("500x600")  # Mở rộng kích thước cửa sổ để chứa bảng
        self.root.resizable(False, False)

        # Nhãn chọn ngôn ngữ
        self.label_lang = tk.Label(root, text="Выберите язык:")
        self.label_lang.pack(pady=5)

        # Dropdown chọn ngôn ngữ
        self.lang_var = tk.StringVar(value="vi-VN")
        self.lang_menu = ttk.Combobox(root, textvariable=self.lang_var, values=["vi-VN", "ru-RU"])
        self.lang_menu.pack(pady=5)

        # Nút nhận diện giọng nói
        self.button_listen = tk.Button(root, text="🎤 Нажмите, чтобы говорить", command=self.recognize_speech)
        self.button_listen.pack(pady=10)

        # Nhãn hiển thị phép tính
        self.label_expression = tk.Label(root, text="Выражение:", font=("Arial", 12))
        self.label_expression.pack(pady=5)

        # Ô hiển thị biểu thức toán học
        self.expression_var = tk.StringVar()
        self.entry_expression = tk.Entry(root, textvariable=self.expression_var, font=("Arial", 14), justify="center")
        self.entry_expression.pack(pady=5, padx=20, fill="x")

        # Nhãn kết quả
        self.label_result = tk.Label(root, text="Результат:", font=("Arial", 12))
        self.label_result.pack(pady=5)

        # Ô hiển thị kết quả
        self.result_var = tk.StringVar()
        self.entry_result = tk.Entry(root, textvariable=self.result_var, font=("Arial", 14), justify="center", state="readonly")
        self.entry_result.pack(pady=5, padx=20, fill="x")

        # Nút lưu kết quả vào Excel
        self.button_save = tk.Button(root, text="💾 Сохранить результаты", command=self.save_to_excel)
        self.button_save.pack(pady=10)

        # Nút xóa phép tính
        self.button_clear = tk.Button(root, text="❌ Ясный расчет", command=self.clear_calculation)
        self.button_clear.pack(pady=10)

        # Nút hủy phép tính
        self.button_cancel = tk.Button(root, text="❌ Отменить расчет", command=self.cancel_calculation)
        self.button_cancel.pack(pady=10)

        # Nút in bảng Excel
        self.button_print_excel = tk.Button(root, text="📄 Распечатать таблицу", command=self.print_excel)
        self.button_print_excel.pack(pady=10)

        # Treeview để hiển thị bảng Excel
        self.treeview = ttk.Treeview(root, columns=("Expression", "Result"), show="headings", height=10)
        self.treeview.heading("Expression", text="Biểu thức")
        self.treeview.heading("Result", text="Kết quả")
        self.treeview.pack(pady=10, padx=20, fill="x")

    def recognize_speech(self):
        recognizer = sr.Recognizer()
        with sr.Microphone() as source:
            self.result_var.set("⏳ Слушание...")
            self.root.update()
            try:
                audio = recognizer.listen(source, timeout=5, phrase_time_limit=5)  # Thiết lập thời gian nghe tối đa cho mỗi câu
                language = self.lang_var.get()
                text = recognizer.recognize_google(audio, language=language).strip().lower()
                self.expression_var.set(self.convert_to_math(text))

                # Tự động tính toán sau khi nghe xong
                expression = self.expression_var.get()
                if expression:
                    result = self.calculate(expression)
                    self.result_var.set(self.format_result(result))

            except sr.UnknownValueError:
                self.result_var.set("❌ Не распознан!")
            except sr.RequestError:
                self.result_var.set("❌ Ошибка соединения!")

    def convert_to_math(self, expression):
        words = expression.split()
        return " ".join([TRANSLATION_DICT.get(word, word) for word in words])

    def calculate(self, expression):
        try:
            return sympify(expression).evalf()
        except Exception as e:
            return f"❌ Ошибка: {e}"

    def format_result(self, result):
        # Sử dụng Decimal để xử lý kết quả và loại bỏ số 0 thừa
        decimal_result = Decimal(str(result))
        if decimal_result == decimal_result.to_integral_value():
            return str(int(decimal_result))  # Chuyển thành kiểu int nếu là số nguyên
        else:
            return str(decimal_result)  # Nếu có phần thập phân thì giữ nguyên

    def save_to_excel(self):
        try:
            # Tạo hoặc mở file Excel
            wb = openpyxl.load_workbook("luu_file.xlsx")
            sheet = wb.active

            # Thêm tiêu đề nếu là lần đầu
            if sheet.max_row == 1:
                sheet.append(["Biểu thức", "Kết quả"])

            # Lưu biểu thức và kết quả vào Excel
            expression = self.expression_var.get()
            result = self.result_var.get()
            sheet.append([expression, result])

            # Lưu file
            wb.save("luu_file.xlsx")
            messagebox.showinfo("Успех", "Результаты сохранены.!")

            # Xóa phép tính
            self.clear_calculation()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Невозможно сохранить : {e}")

    def clear_calculation(self):
        # Xóa phép tính và kết quả
        self.expression_var.set("")
        self.result_var.set("")

    def cancel_calculation(self):
        # Xóa phép tính và kết quả mà không lưu
        self.clear_calculation()
        messagebox.showinfo("Отменить расчет", "Расчет отменен.")

    def print_excel(self):
        # Đọc và hiển thị dữ liệu từ file Excel vào Treeview
        try:
            wb = openpyxl.load_workbook("luu_file.xlsx")
            sheet = wb.active

            # Xóa các dòng cũ trong treeview
            for row in self.treeview.get_children():
                self.treeview.delete(row)

            # Đọc dữ liệu từ sheet và thêm vào treeview
            for row in sheet.iter_rows(min_row=2, values_only=True):  # Bỏ qua dòng tiêu đề
                self.treeview.insert("", "end", values=row)

        except Exception as e:
            messagebox.showerror("Ошибка", f"Невозможно сохранить : {e}")
    
    def format_result(self, result):
        try:
            # Kiểm tra xem kết quả có phải là số hợp lệ không
            decimal_result = Decimal(str(result))
            # Nếu là số nguyên, chuyển thành int
            if decimal_result == decimal_result.to_integral_value():
                return str(int(decimal_result))  # Chuyển thành kiểu int nếu là số nguyên
            else:
                return str(decimal_result)  # Nếu có phần thập phân thì giữ nguyên
        except (ValueError, InvalidOperation):
            # Trường hợp nếu kết quả không hợp lệ
            return "❌ Lỗi: Kết quả không hợp lệ!"


if __name__ == "__main__":
    root = tk.Tk()
    app = VoiceCalculatorApp(root)
    root.mainloop()
